import os
#  from string import ascii_uppercase
from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelManager:
    def __init__(self, outputFilename: str = "output.xlsx", useBoldTitles: bool = True, saveOnFill: bool = False,
                 resizeToFit: bool = False):
        self.workBook = Workbook()  # Объект Excel'я
        self.sheet = self.workBook.active  # Сама страница
        self.useBoldTitles = useBoldTitles  # Сделать заголовки жирным шрифтом
        self.resizeToFit = resizeToFit  # Автоматически расширить столбцы
        self.saveOnFill = saveOnFill  # Автоматическая перезапись файла при еще одном заполнении
        self.outputFilename = outputFilename  # Название выходного файла
        self.setOutputFilename(outputFilename)

    def fillSheet(self, titles: list, data: list):
        """
        Заполнение Excel файла
        :param titles: Список заголовков
        :param data: Список данных, которые должны стоять под заголовками
        :return: Заполняет внутреннюю переменную self.sheet
        """

        """
        Структура входных данных
        
        titles: list = [
            'title1',
            'title2',
            ...
        ]
        
        data: list = [
            [
                'item1',
                'item2',
                'item3',
                'item4',
                ...
            ],
            [
                'item1',
                'item2',
                'item3',
                'item4',
                ...
            ],
            ...
        ]
        """

        # Определение максимального алфавита столбцов
        # alphabet = ascii_uppercase[:len(titles)]

        # Заполнение заголовков
        for h, title in enumerate(titles):
            cell = self.sheet.cell(row=1, column=h + 1)
            cell.value = title

            # Заполняем заголовки жирным шрифтом
            if self.useBoldTitles:
                cell.font = Font(bold=True)

        # Заполнение строчек данными
        for h, item in enumerate(data):
            for q, subItem in enumerate(item):
                cell = self.sheet.cell(row=h + 2, column=q + 1)
                if isinstance(subItem, list) or isinstance(subItem, tuple) or isinstance(subItem, set):
                    cell.value = self.convertArrayToString(subItem)
                elif isinstance(subItem, str):
                    cell.value = subItem
                elif isinstance(subItem, int):
                    cell.value = str(subItem)
                else:
                    raise Exception(f"Получен неизвестный (необработанный)"
                                    f" тип элемента при записи данных. ({type(subItem)})")

        if self.resizeToFit:
            self.resize()

        if self.saveOnFill:
            self.save()

    def resize(self):
        """
        Автоматическое регулирование размеров столбцов для полного вмещения в них контента
        :return:
        """
        dims = {}
        for row in self.sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            self.sheet.column_dimensions[col].width = value

    def save(self):
        """
        Сохранение Excel файла
        :return:
        """
        while True:
            try:
                self.workBook.save(self.outputFilename)
                print(f"Файл \"{self.outputFilename}\" сохранен успешно!")
                break
            except PermissionError:
                print("Произошла ошибка сохранения Excel файла!\n"
                      "Попробуйте закрыть файл, затем повторите попытку.")
                input("[X] Нажмите Enter чтобы повторить попытку..")

    def convertArrayToString(self, obj, separator: str = ", ", end: str = "."):
        """
        Форматирование массива в строку
        :param obj: Массив, который необходимо конвертировать (list, tuple,
        :param separator: Разделитель между элементами списка
        :param end: Конечный символ
        :return: Форматированная строка
        """
        return separator.join(obj)[:-1] + end if len(obj) != 0 else "-"

    def setOutputFilename(self, filename: str):
        """
        Проверка правильности ввода имени выходного файла.
        :param filename: Имя файла.
        :return: Если данные введены корректно - имя выходного файла.
        """
        newName = False
        if isinstance(filename, str):
            while True:
                if newName:
                    filename = input("Выберите новое имя файла: ")
                    newName = False
                if filename.endswith(".xlsx") and len(filename.strip()) > 5:
                    if os.path.isfile(filename):
                        print(f"[%] Файл с таким названием ({self.outputFilename}) уже существует.")
                        rewrite = input("Перезаписать? (+/-): ")
                        if rewrite == "+":
                            print("[*] Файл будет перезаписан.")
                            while True:
                                try:
                                    os.remove(filename)
                                    break
                                except PermissionError:
                                    print("[!] Возникла ошибка!\n"
                                          "Необходимо закрыть этот Excel документ перед его перезаписыванием.")
                                    input("Нажмите Enter чтобы попробовать еще раз..")
                            self.outputFilename = filename
                            break
                        elif rewrite == "-":
                            newName = True
                        else:
                            print("[!] Команда не распознана.")
                    else:
                        break
                else:
                    print("[!] Неверно указано название файла. Пример: out.xlsx")
                    newName = True

            self.outputFilename = filename
            print("[*] Выходной файл установлен успешно. ")


if __name__ == '__main__':
    # Пример использования
    excelMan = ExcelManager()
    titles = ['title1', 'title2']
    data = [['crawledData1', ['item1', 'item2', 'item3']], ['crawledData1', 'crawledData2']]

    excelMan.fillSheet(titles, data)
    excelMan.save()
